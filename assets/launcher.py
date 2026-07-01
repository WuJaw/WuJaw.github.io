import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import subprocess
import threading
import os
import sys
import json
import queue
from datetime import datetime

try:
    import serial.tools.list_ports
    HAS_PYSERIAL = True
except ImportError:
    HAS_PYSERIAL = False

try:
    from openpyxl import load_workbook, Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─────────────────────────────────────────────
# 配置：可执行文件列表
# ─────────────────────────────────────────────
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "launcher_config.json")

DEFAULT_CONFIG = {
    "port": "",
    "excel": "请先选择文件",
    "excel_fullpath": "",
    "ble_tag": "56 00 06 32",
    "rssi_thresh": "-50",
    "tag_rssi_thresh": "-30",
    "station_id": "69 00 26 40",
    "ver_master": "00 01 00 00",
    "ver_slave1": "00 01 00 00",
    "ver_slave2": "00 07 00 00",
    "ble_packet_count": "10",
    "ble_timeout": "20",
}

PROGRAMS = [
    {
        "label": "正式版",
        "exe": os.path.join(BASE_DIR, "DataMicrostationRssiTest.exe"),
        "desc": "DataMicrostationRssiTest.exe",
    },
    {
        "label": "半成品测试版",
        "exe": os.path.join(BASE_DIR, "DataMicrostationRssiTest半成品测试.exe"),
        "desc": "DataMicrostationRssiTest半成品测试.exe",
    },
]


# ─────────────────────────────────────────────
# 颜色主题
# ─────────────────────────────────────────────
BG        = "#F5F6FA"
PANEL_BG  = "#FFFFFF"
ACCENT    = "#4A6CF7"
ACCENT_HV = "#3A5CE7"
BTN_STOP  = "#E74C3C"
BTN_STOP_HV = "#C0392B"
TEXT_MAIN = "#1A1A2E"
TEXT_SUB  = "#666680"
LOG_BG    = "#1E1E2E"
LOG_FG    = "#CDD6F4"
LOG_INFO  = "#89DCEB"
LOG_OK    = "#A6E3A1"
LOG_ERR   = "#F38BA8"
LOG_WARN  = "#FAB387"
BORDER    = "#E0E0F0"
TEXT_OK   = "#27AE60"   # 绿色 - 版本正常
TEXT_ERR  = "#E53935"   # 红色 - 版本异常/超时


class LauncherApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("数分微站专用测试软件")

        # 自适应屏幕高度：低分辨率屏幕（如 1366×768）自动缩小窗口
        screen_h = self.winfo_screenheight()
        win_h = min(680, max(580, screen_h - 90))
        win_w = min(975, self.winfo_screenwidth() - 40)
        self.geometry(f"{win_w}x{win_h}")
        self.resizable(False, False)
        self.configure(bg=BG)

        # 加载配置
        self._cfg = self._load_config()

        # 运行状态
        self.proc = None
        self.log_queue = queue.Queue()
        self.running = False
        self.selected_program = None
        self._config_widgets = []   # 收集所有可禁用的配置控件
        self._save_pending = False  # 防抖标记
        self._test_mode = ""         # "" | "full" | "semi"

        self._build_ui()
        self._bind_autosave()
        self._poll_log()

    # ─────────────────────────────────────────
    # 配置持久化
    # ─────────────────────────────────────────
    def _load_config(self):
        """从 JSON 读取配置，不存在则返回默认值"""
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    saved = json.load(f)
                cfg = DEFAULT_CONFIG.copy()
                cfg.update(saved)
                return cfg
            except Exception:
                return DEFAULT_CONFIG.copy()
        return DEFAULT_CONFIG.copy()

    def _save_config(self, *_):
        """自动保存（带防抖）"""
        if self._save_pending:
            return
        self._save_pending = True
        self.after(300, self._do_save_config)

    def _do_save_config(self):
        self._save_pending = False
        data = {
            "port": self.var_port.get(),
            "excel": self.var_excel.get(),
            "excel_fullpath": getattr(self, "_excel_fullpath", self._cfg.get("excel_fullpath", "")),
            "ble_tag": self.var_ble_tag.get(),
            "rssi_thresh": self.var_rssi_thresh.get(),
            "tag_rssi_thresh": self.var_tag_rssi_thresh.get(),
            "station_id": self.var_station_id.get(),
            "ver_master": self.var_ver_master.get(),
            "ver_slave1": self.var_ver_slave1.get(),
            "ver_slave2": self.var_ver_slave2.get(),
            "ble_packet_count": self.var_ble_packet_count.get(),
            "ble_timeout": self.var_ble_timeout.get(),
        }
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _bind_autosave(self):
        """绑定所有 StringVar 的 change 事件 → 自动保存"""
        vars_to_trace = [
            self.var_port, self.var_excel, self.var_ble_tag,
            self.var_rssi_thresh, self.var_tag_rssi_thresh, self.var_station_id,
            self.var_ver_master, self.var_ver_slave1, self.var_ver_slave2,
            self.var_ble_packet_count,
            self.var_ble_timeout,
        ]
        for v in vars_to_trace:
            v.trace_add("write", self._save_config)

    # ─────────────────────────────────────────
    # 构建 UI
    # ─────────────────────────────────────────
    def _build_ui(self):
        # 主体容器
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=14, pady=12)

        # 左侧面板
        left = tk.Frame(body, bg=PANEL_BG, bd=0, highlightthickness=1,
                        highlightbackground=BORDER, width=260)
        left.pack(side="left", fill="y", padx=(0, 6))
        left.pack_propagate(False)

        self._build_left(left)

        # 右侧日志面板
        right = tk.Frame(body, bg=PANEL_BG, bd=0, highlightthickness=1,
                         highlightbackground=BORDER)
        right.pack(side="left", fill="both", expand=True)

        self._build_right(right)

    def _build_left(self, parent):
        # ── 串口号选择 ──
        port_row = tk.Frame(parent, bg=PANEL_BG)
        port_row.pack(fill="x", padx=4, pady=(10, 0))
        tk.Label(
            port_row, text="串口号", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="w", width=10,
        ).pack(side="left")

        self.var_port = tk.StringVar(value=self._cfg["port"])
        self.cmb_port = ttk.Combobox(
            port_row,
            textvariable=self.var_port,
            state="readonly",
            font=("Microsoft YaHei UI", 10),
        )
        self.cmb_port.pack(side="left", fill="x", expand=True)
        self._config_widgets.append(self.cmb_port)

        self._scan_ports()

        # ── Excel 配置文件选择 ──
        excel_row = tk.Frame(parent, bg=PANEL_BG)
        excel_row.pack(fill="x", padx=4, pady=(6, 0))
        tk.Label(
            excel_row, text="写入文件", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="w", width=10,
        ).pack(side="left")

        self.var_excel = tk.StringVar(value=self._cfg["excel"])
        self._excel_fullpath = self._cfg.get("excel_fullpath", "")
        self.ent_excel = tk.Entry(
            excel_row,
            textvariable=self.var_excel,
            state="readonly",
            readonlybackground=PANEL_BG,
            relief="solid",
            bd=1,
            font=("Microsoft YaHei UI", 9),
            fg=TEXT_MAIN,
        )
        self.ent_excel.pack(side="left", fill="x", expand=True, ipady=2)

        btn_excel = self._make_button(
            excel_row, "…", ACCENT, ACCENT_HV, self._pick_excel,
            height=22, font_size=10
        )
        btn_excel.pack(side="left", padx=(4, 0))
        self.btn_excel = btn_excel
        self._config_widgets.append(self.ent_excel)
        self._config_widgets.append(btn_excel)

        # ── 蓝牙测试标签 ──
        ble_row = tk.Frame(parent, bg=PANEL_BG)
        ble_row.pack(fill="x", padx=4, pady=(4, 0))
        tk.Label(
            ble_row, text="蓝牙测试标签", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="w", width=10,
        ).pack(side="left")

        self.var_ble_tag = tk.StringVar(value=self._cfg["ble_tag"])
        self.ent_ble_tag = tk.Entry(
            ble_row,
            textvariable=self.var_ble_tag,
            relief="solid",
            bd=1,
            font=("Microsoft YaHei UI", 10),
            fg=TEXT_MAIN,
        )
        self.ent_ble_tag.pack(side="left", fill="x", expand=True, ipady=2)
        self.ent_ble_tag.bind("<FocusOut>", lambda e, v=self.var_ble_tag: self._on_hex_entry_focus_out(v, e))
        self._config_widgets.append(self.ent_ble_tag)

        # ── 基站信号强度阈值 ──
        rssi_row = tk.Frame(parent, bg=PANEL_BG)
        rssi_row.pack(fill="x", padx=4, pady=(6, 0))
        tk.Label(
            rssi_row, text="基站信号阈值", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="w", width=10,
        ).pack(side="left")

        self.var_rssi_thresh = tk.StringVar(value=self._cfg["rssi_thresh"])
        self.ent_rssi_thresh = tk.Entry(
            rssi_row,
            textvariable=self.var_rssi_thresh,
            relief="solid",
            bd=1,
            font=("Microsoft YaHei UI", 10),
            fg=TEXT_MAIN,
        )
        self.ent_rssi_thresh.pack(side="left", fill="x", expand=True, ipady=2)
        self._config_widgets.append(self.ent_rssi_thresh)

        # ── 标签信号强度阈值 ──
        tag_rssi_row = tk.Frame(parent, bg=PANEL_BG)
        tag_rssi_row.pack(fill="x", padx=4, pady=(6, 0))
        tk.Label(
            tag_rssi_row, text="标签信号强度阈值", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="w", width=10,
        ).pack(side="left")

        self.var_tag_rssi_thresh = tk.StringVar(value=self._cfg["tag_rssi_thresh"])
        self.ent_tag_rssi_thresh = tk.Entry(
            tag_rssi_row,
            textvariable=self.var_tag_rssi_thresh,
            relief="solid",
            bd=1,
            font=("Microsoft YaHei UI", 10),
            fg=TEXT_MAIN,
        )
        self.ent_tag_rssi_thresh.pack(side="left", fill="x", expand=True, ipady=2)
        self._config_widgets.append(self.ent_tag_rssi_thresh)

        # ── 信号采集包数 ──
        pkt_row = tk.Frame(parent, bg=PANEL_BG)
        pkt_row.pack(fill="x", padx=4, pady=(6, 0))
        tk.Label(
            pkt_row, text="信号采集包数", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="w", width=10,
        ).pack(side="left")

        self.var_ble_packet_count = tk.StringVar(value=self._cfg["ble_packet_count"])
        self.ent_ble_packet_count = tk.Entry(
            pkt_row,
            textvariable=self.var_ble_packet_count,
            relief="solid",
            bd=1,
            font=("Microsoft YaHei UI", 10),
            fg=TEXT_MAIN,
        )
        self.ent_ble_packet_count.pack(side="left", fill="x", expand=True, ipady=2)
        self._config_widgets.append(self.ent_ble_packet_count)

        # ── 超时时间 ──
        tmt_row = tk.Frame(parent, bg=PANEL_BG)
        tmt_row.pack(fill="x", padx=4, pady=(6, 0))
        tk.Label(
            tmt_row, text="超时时间(秒)", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="w", width=10,
        ).pack(side="left")

        self.var_ble_timeout = tk.StringVar(value=self._cfg["ble_timeout"])
        self.ent_ble_timeout = tk.Entry(
            tmt_row,
            textvariable=self.var_ble_timeout,
            relief="solid",
            bd=1,
            font=("Microsoft YaHei UI", 10),
            fg=TEXT_MAIN,
        )
        self.ent_ble_timeout.pack(side="left", fill="x", expand=True, ipady=2)
        self._config_widgets.append(self.ent_ble_timeout)

        # ── 主机版本00 ──
        ver1_row = tk.Frame(parent, bg=PANEL_BG)
        ver1_row.pack(fill="x", padx=4, pady=(6, 0))
        tk.Label(
            ver1_row, text="主机版本00", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="w", width=10,
        ).pack(side="left")

        self.var_ver_master = tk.StringVar(value=self._cfg["ver_master"])
        self.ent_ver_master = tk.Entry(
            ver1_row,
            textvariable=self.var_ver_master,
            relief="solid",
            bd=1,
            font=("Microsoft YaHei UI", 10),
            fg=TEXT_MAIN,
        )
        self.ent_ver_master.pack(side="left", fill="x", expand=True, ipady=2)
        self.ent_ver_master.bind("<FocusOut>", lambda e, v=self.var_ver_master: self._on_hex_entry_focus_out(v, e))
        self._config_widgets.append(self.ent_ver_master)

        # ── 从机版本01 ──
        ver2_row = tk.Frame(parent, bg=PANEL_BG)
        ver2_row.pack(fill="x", padx=4, pady=(6, 0))
        tk.Label(
            ver2_row, text="从机版本 01", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="w", width=10,
        ).pack(side="left")

        self.var_ver_slave1 = tk.StringVar(value=self._cfg["ver_slave1"])
        self.ent_ver_slave1 = tk.Entry(
            ver2_row,
            textvariable=self.var_ver_slave1,
            relief="solid",
            bd=1,
            font=("Microsoft YaHei UI", 10),
            fg=TEXT_MAIN,
        )
        self.ent_ver_slave1.pack(side="left", fill="x", expand=True, ipady=2)
        self.ent_ver_slave1.bind("<FocusOut>", lambda e, v=self.var_ver_slave1: self._on_hex_entry_focus_out(v, e))
        self._config_widgets.append(self.ent_ver_slave1)

        # ── 从机版本02 ──
        ver3_row = tk.Frame(parent, bg=PANEL_BG)
        ver3_row.pack(fill="x", padx=4, pady=(6, 0))
        tk.Label(
            ver3_row, text="从机版本02", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="w", width=10,
        ).pack(side="left")

        self.var_ver_slave2 = tk.StringVar(value=self._cfg["ver_slave2"])
        self.ent_ver_slave2 = tk.Entry(
            ver3_row,
            textvariable=self.var_ver_slave2,
            relief="solid",
            bd=1,
            font=("Microsoft YaHei UI", 10),
            fg=TEXT_MAIN,
        )
        self.ent_ver_slave2.pack(side="left", fill="x", expand=True, ipady=2)
        self.ent_ver_slave2.bind("<FocusOut>", lambda e, v=self.var_ver_slave2: self._on_hex_entry_focus_out(v, e))
        self._config_widgets.append(self.ent_ver_slave2)

        # ── 数分微站号 ──
        station_row = tk.Frame(parent, bg=PANEL_BG)
        station_row.pack(fill="x", padx=4, pady=(6, 0))
        tk.Label(
            station_row, text="数分微站号", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="w", width=10,
        ).pack(side="left")

        self.var_station_id = tk.StringVar(value=self._cfg["station_id"])
        self.ent_station_id = tk.Entry(
            station_row,
            textvariable=self.var_station_id,
            relief="solid",
            bd=1,
            font=("Microsoft YaHei UI", 10),
            fg=TEXT_MAIN,
        )
        self.ent_station_id.pack(side="left", fill="x", expand=True, ipady=2)
        self.ent_station_id.bind("<FocusOut>", lambda e, v=self.var_station_id: self._on_hex_entry_focus_out(v, e))
        self._config_widgets.append(self.ent_station_id)

        # ── 操作按钮 ──
        btn_row = tk.Frame(parent, bg=PANEL_BG)
        btn_row.pack(fill="x", padx=4, pady=(8, 0))

        self.btn_full = self._make_button(
            btn_row, "成品测试", ACCENT, ACCENT_HV, self._start_full_test,
        )
        self.btn_full.pack(side="left", fill="x", expand=True, padx=(0, 4))

        self.btn_semi = self._make_button(
            btn_row, "半成品测试", ACCENT, ACCENT_HV, self._start_semi_test,
        )
        self.btn_semi.pack(side="left", fill="x", expand=True, padx=(0, 4))

        self.btn_stop = self._make_button(
            btn_row, "停止测试", BTN_STOP, BTN_STOP_HV, self._stop,
        )
        self.btn_stop.pack(side="left", fill="x", expand=True)

        # 将所有配置控件加入锁控列表
        self._config_widgets.append(self.btn_full)
        self._config_widgets.append(self.btn_semi)

        # ── 重试写入按钮（写入失败时显示）──
        self.btn_retry = self._make_button(
            parent, "重新写入", "#D35400", "#E67E22", self._retry_write,
        )
        self.btn_retry.pack_forget()  # 默认隐藏
        self._pending_write_data = None

        # ── 执行进度条 ──
        self.progress_frame = tk.Frame(parent, bg=PANEL_BG)
        self.progress_frame.pack(fill="x", padx=4, pady=(10, 2))

        self.var_progress_pct = tk.StringVar(value="0%")
        self.progress_bar = ttk.Progressbar(
            self.progress_frame,
            mode="determinate",
            maximum=100,
            value=0,
        )
        self.progress_bar.pack(side="left", fill="x", expand=True)

        self.lbl_progress = tk.Label(
            self.progress_frame, textvariable=self.var_progress_pct,
            bg=PANEL_BG, fg=TEXT_SUB,
            font=("Microsoft YaHei UI", 9),
            width=5, anchor="e",
        )
        self.lbl_progress.pack(side="left", padx=(4, 0))

        self.lbl_progress_time = tk.Label(
            self.progress_frame, text="",
            bg=PANEL_BG, fg=TEXT_SUB,
            font=("Microsoft YaHei UI", 8),
        )
        self.lbl_progress_time.pack(side="left", padx=(6, 0))

        # ── 实时信号强度显示 ──
        sig_sep = tk.Frame(parent, bg=BORDER, height=1)
        sig_sep.pack(fill="x", padx=4, pady=(4, 2))

        # 与基站信号
        row_rssi = tk.Frame(parent, bg=PANEL_BG)
        row_rssi.pack(fill="x", padx=4, pady=(1, 0))
        tk.Label(
            row_rssi, text="与基站信号", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="e", width=12,
        ).pack(side="left")
        self.var_rssi_disp = tk.StringVar(value="— dBm")
        tk.Label(
            row_rssi, textvariable=self.var_rssi_disp, bg=PANEL_BG,
            fg=TEXT_MAIN, font=("Microsoft YaHei UI", 9, "bold"),
            anchor="w",
        ).pack(side="left", padx=(4, 0))
        self.var_rssi_count = tk.StringVar(value="")
        tk.Label(
            row_rssi, textvariable=self.var_rssi_count, bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 8),
            anchor="e",
        ).pack(side="right")

        # 与蓝牙标签信号
        row_ble = tk.Frame(parent, bg=PANEL_BG)
        row_ble.pack(fill="x", padx=4, pady=(1, 0))
        self._row_ble = row_ble  # 保存引用，供版本面板插入位置使用
        tk.Label(
            row_ble, text="与蓝牙标签信号", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
            anchor="e", width=12,
        ).pack(side="left")
        self.var_ble_rssi_disp = tk.StringVar(value="— dBm")
        tk.Label(
            row_ble, textvariable=self.var_ble_rssi_disp, bg=PANEL_BG,
            fg=TEXT_MAIN, font=("Microsoft YaHei UI", 9, "bold"),
            anchor="w",
        ).pack(side="left", padx=(4, 0))
        self.var_ble_rssi_count = tk.StringVar(value="")
        tk.Label(
            row_ble, textvariable=self.var_ble_rssi_count, bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 8),
            anchor="e",
        ).pack(side="right")

        # 警告变量（控件在底部与提示一起）
        self.var_rssi_warn = tk.StringVar(value="")
        self.var_ble_rssi_warn = tk.StringVar(value="")

        # ── 设备版本信息（仅半成品测试显示） ──
        self.frame_version = tk.Frame(parent, bg=PANEL_BG)

        self._ver_labels = {}
        self._ver_vars = {}
        for ver_name in ["主机版本00", "从机版本01", "从机版本02"]:
            row = tk.Frame(self.frame_version, bg=PANEL_BG)
            row.pack(fill="x", padx=4, pady=(0, 0))
            tk.Label(
                row, text=ver_name, bg=PANEL_BG,
                fg=TEXT_SUB, font=("Microsoft YaHei UI", 9),
                anchor="e", width=10,
            ).pack(side="left")
            sv = tk.StringVar(value="—")
            self._ver_vars[ver_name] = sv
            self._ver_labels[ver_name] = tk.Label(
                row, textvariable=sv, bg=PANEL_BG,
                fg=ACCENT, font=("Consolas", 12, "bold"),
                anchor="e",
            )
            self._ver_labels[ver_name].pack(side="left")

        # 版本状态提示
        self.var_ver_status = tk.StringVar(value="")
        self.lbl_ver_status = tk.Label(
            self.frame_version, textvariable=self.var_ver_status, bg=PANEL_BG,
            font=("Microsoft YaHei UI", 10, "bold"), anchor="w",
        )
        self.lbl_ver_status.pack(fill="x", padx=4, pady=(6, 2))

        # 版本面板先占位，默认隐藏
        self.frame_version.pack(fill="x", padx=4, pady=(0, 0))
        self.frame_version.pack_forget()

        # ── 信号/提示行（警告与提示一起放在最底部）──
        hint_row = tk.Frame(parent, bg=PANEL_BG)
        hint_row.pack(fill="x", padx=4, pady=(4, 2))

        tk.Label(
            hint_row, textvariable=self.var_rssi_warn, bg=PANEL_BG,
            fg=TEXT_ERR, font=("Microsoft YaHei UI", 8, "bold"),
            anchor="w",
        ).pack(side="left", fill="x", expand=True)

        tk.Label(
            hint_row, textvariable=self.var_ble_rssi_warn, bg=PANEL_BG,
            fg=TEXT_ERR, font=("Microsoft YaHei UI", 8, "bold"),
            anchor="w",
        ).pack(side="left", fill="x", expand=True)

        # 底部版本标记
        self._bottom_label = tk.Label(
            parent, text="", bg=PANEL_BG,
            fg="#BBBBCC", font=("Microsoft YaHei UI", 8)
        )
        self._bottom_label.pack(side="bottom")

    def _build_right(self, parent):
        header = tk.Frame(parent, bg=PANEL_BG)
        header.pack(fill="x", padx=14, pady=(12, 4))
        tk.Label(
            header, text="实时日志", bg=PANEL_BG,
            fg=TEXT_MAIN, font=("Microsoft YaHei UI", 11, "bold")
        ).pack(side="left")
        self.lbl_log_hint = tk.Label(
            header, text="等待启动…", bg=PANEL_BG,
            fg=TEXT_SUB, font=("Microsoft YaHei UI", 9)
        )
        self.lbl_log_hint.pack(side="right")

        sep = tk.Frame(parent, bg=BORDER, height=1)
        sep.pack(fill="x", padx=14, pady=(0, 8))

        self.log_box = scrolledtext.ScrolledText(
            parent,
            bg=LOG_BG, fg=LOG_FG,
            insertbackground=LOG_FG,
            font=("Consolas", 9),
            bd=0,
            relief="flat",
            state="disabled",
            wrap="word",
        )
        self.log_box.pack(fill="both", expand=True, padx=14, pady=(0, 12))

        # 颜色 tag
        self.log_box.tag_config("info",  foreground=LOG_INFO)
        self.log_box.tag_config("ok",    foreground=LOG_OK)
        self.log_box.tag_config("err",   foreground=LOG_ERR)
        self.log_box.tag_config("warn",  foreground=LOG_WARN)
        self.log_box.tag_config("plain", foreground=LOG_FG)
        self.log_box.tag_config("ts",    foreground="#6C7086")
        self.log_box.tag_config("match_ble",     foreground="#FFFFFF",
                                background="#8E24AA")  # 紫底白字 - 蓝牙标签
        self.log_box.tag_config("match_station", foreground="#FFFFFF",
                                background="#43A047")  # 绿底白字 - 数分微站号
        self.log_box.tag_config("ble_7th", foreground="#FFFFFF",
                                background="#E53935")  # 红底白字 - 蓝牙匹配后第7字节
        self.log_box.tag_config("last",  foreground="#333333",
                                background="#FFF176")  # 黄底黑字 - 最后字节

    # ─────────────────────────────────────────
    # 辅助构建
    # ─────────────────────────────────────────
    def _make_button(self, parent, text, color, hover_color, command,
                     height=26, font_size=9):
        btn = tk.Label(
            parent, text=text, bg=color, fg="white",
            font=("Microsoft YaHei UI", font_size, "bold"),
            height=1, cursor="hand2", pady=6,
        )
        btn.bind("<Button-1>", lambda e: command() if str(btn["state"]) != "disabled" else None)
        btn.bind("<Enter>", lambda e: btn.config(bg=hover_color) if str(btn["state"]) != "disabled" else None)
        btn.bind("<Leave>", lambda e: btn.config(bg=color) if str(btn["state"]) != "disabled" else None)
        btn._color = color
        btn._hover = hover_color

        def _set_state(state):
            btn.config(state=state)
            if state == "disabled":
                btn.config(bg="#CCCCCC")
                btn.unbind("<Enter>")
                btn.unbind("<Leave>")
            else:
                btn.config(bg=color)
                btn.bind("<Enter>", lambda e: btn.config(bg=hover_color))
                btn.bind("<Leave>", lambda e: btn.config(bg=color))

        btn.config = btn.config  # keep original
        btn._set_state = _set_state
        return btn

    # ─────────────────────────────────────────
    # 串口扫描
    # ─────────────────────────────────────────
    def _scan_ports(self):
        ports = []
        if HAS_PYSERIAL:
            for p in serial.tools.list_ports.comports():
                ports.append(p.device)
        else:
            # 回退：尝试 COM1 ~ COM16
            for i in range(1, 17):
                ports.append(f"COM{i}")

        current = self.var_port.get()
        self.cmb_port["values"] = ports
        if current in ports:
            self.var_port.set(current)
        elif ports:
            self.var_port.set(ports[0])
        else:
            self.var_port.set("")

    # ─────────────────────────────────────────
    # Excel 文件选择
    # ─────────────────────────────────────────
    def _pick_excel(self):
        path = filedialog.askopenfilename(
            title="选择 Excel 配置文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
            initialdir=BASE_DIR,
        )
        if path:
            self._excel_fullpath = path
            self.var_excel.set(os.path.basename(path))
            self._log(f"已选择 Excel：{path}", "info")

    # ─────────────────────────────────────────
    # 配置锁定/解锁
    # ─────────────────────────────────────────
    def _lock_config(self):
        """测试开始：禁用所有配置控件"""
        for w in self._config_widgets:
            if isinstance(w, ttk.Combobox):
                w.config(state="disabled")
            elif hasattr(w, "_set_state"):
                w._set_state("disabled")
            else:
                w.config(state="disabled")

    def _unlock_config(self):
        """测试结束：恢复所有配置控件"""
        for w in self._config_widgets:
            if isinstance(w, ttk.Combobox):
                w.config(state="readonly")
            elif hasattr(w, "_set_state"):
                w._set_state("normal")
            else:
                w.config(state="normal")

    def _show_version_panel(self):
        """显示版本信息面板（半成品测试专用）"""
        self.frame_version.pack(fill="x", padx=4, pady=(0, 0), after=self._row_ble)
        self._collected_versions = {}  # name → ver_str
        self._versions_checked = False
        # 重置信号计数器 + 值列表
        self._rssi_count = 0; self._ble_rssi_count = 0
        self._rssi_values = []; self._ble_rssi_values = []
        self.var_rssi_count.set(""); self.var_ble_rssi_count.set("")
        self.var_rssi_warn.set(""); self.var_ble_rssi_warn.set("")
        for sv in self._ver_vars.values():
            sv.set("—")
        self.var_ver_status.set("")

    def _hide_version_panel(self):
        """隐藏版本信息面板"""
        self.frame_version.pack_forget()
        for sv in self._ver_vars.values():
            sv.set("—")
        self.var_ver_status.set("")

    def _check_versions_complete(self):
        """版本全部采集完成时：停进度条 + 对比配置版本"""
        if self._versions_checked:
            return
        station = self.var_station_id.get().replace(" ", "")
        # 65 开头只有主机+从机01，其余有 3 个版本
        required = {"主机版本00", "从机版本01"} if station[:2] == "65" else \
                   {"主机版本00", "从机版本01", "从机版本02"}

        if not required.issubset(self._collected_versions.keys()):
            return

        self._versions_checked = True
        # 只停定时器，不 reset — 让进度条停在 100%
        pid = getattr(self, "_progress_id", None)
        if pid is not None:
            self.after_cancel(pid)
            self._progress_id = None
        self.progress_bar["value"] = 100
        self.var_progress_pct.set("100%")
        self._log("版本采集完成", "ok")
        # 采集完成 → 关闭串口
        self.running = False

        # ── 信号中位数计算与阈值比较 ──
        self._check_rssi_thresholds()

        config_map = {
            "主机版本00": self.var_ver_master.get(),
            "从机版本01": self.var_ver_slave1.get(),
            "从机版本02": self.var_ver_slave2.get(),
        }
        all_match = True
        for name in required:
            collected = self._collected_versions.get(name, "")
            expected = config_map[name]
            if collected.replace(" ", "").upper() != expected.replace(" ", "").upper():
                all_match = False
                self._log(f"{name} 不匹配：采集 {collected} ≠ 配置 {expected}", "err")

        if all_match:
            self.var_ver_status.set("测试成功，版本正常")
            self.lbl_ver_status.config(fg=TEXT_OK)
            self._log("测试成功，版本正常", "ok")
        else:
            self.var_ver_status.set("版本异常")
            self.lbl_ver_status.config(fg=TEXT_ERR)
            self._log("版本校验未通过", "warn")

        # 写入 Excel（半成品测试无手动蓝牙输入）
        self._write_to_excel()
        # 解锁配置，保留信号和版本显示
        self._unlock_config()

    # ─────────────────────────────────────────
    # 事件回调
    # ─────────────────────────────────────────
    def _on_select(self):
        idx = self.var_prog.get()
        prog = PROGRAMS[idx]
        self.selected_program = prog
        self.lbl_path.config(text=prog["desc"])

    def _start_full_test(self):
        """成品测试：打开串口 → 开始采集信号 → 完成后弹出手机蓝牙输入 → 写入Excel → 结束"""
        if self.running:
            self._log("已有任务在运行，请等待结束", "warn")
            return

        port = self.var_port.get()
        if not port:
            messagebox.showwarning("未选择串口", "请先在配置中选择串口号")
            return

        excel_path = getattr(self, "_excel_fullpath", "")
        if not excel_path or not os.path.isfile(excel_path):
            messagebox.showwarning("未选择写入文件", "请先在配置中选择写入文件（Excel）")
            return

        self._test_mode = "full"
        self.running = True
        self._lock_config()
        self._pending_write_data = None
        self.btn_retry.pack_forget()
        self._hide_version_panel()    # 隐藏半成品版本面板
        # 重置信号显示
        self.var_rssi_disp.set("— dBm"); self.var_ble_rssi_disp.set("— dBm")
        self._rssi_count = 0; self._ble_rssi_count = 0
        self._rssi_values = []; self._ble_rssi_values = []
        self.var_rssi_count.set(""); self.var_ble_rssi_count.set("")
        self.var_rssi_warn.set(""); self.var_ble_rssi_warn.set("")
        self._start_progress()
        self.lbl_log_hint.config(text=f"成品测试 → {port}")
        self._log(f"打开串口 {port}，开始采集信号强度…", "info")
        # 启动超时定时器
        try:
            timeout_sec = int(self.var_ble_timeout.get())
        except (ValueError, TypeError):
            timeout_sec = 20
        self._full_test_timeout_id = self.after(timeout_sec * 1000, self._on_full_test_timeout)

        threading.Thread(target=self._run_serial, args=(port,), daemon=True).start()

    def _start_semi_test(self):
        """启动半成品测试：打开串口直连测试"""
        if self.running:
            self._log("已有任务在运行，请等待结束", "warn")
            return

        port = self.var_port.get()
        if not port:
            messagebox.showwarning("未选择串口", "请先在配置中选择串口号")
            return

        self._test_mode = "semi"
        self.running = True
        self._lock_config()
        self._pending_write_data = None
        self.btn_retry.pack_forget()
        self._start_progress(self.SEMI_DURATION)
        # 重置信号显示
        self.var_rssi_disp.set("— dBm"); self.var_ble_rssi_disp.set("— dBm")
        self._rssi_count = 0; self._ble_rssi_count = 0
        self._rssi_values = []; self._ble_rssi_values = []
        self.var_rssi_count.set(""); self.var_ble_rssi_count.set("")
        self.var_rssi_warn.set(""); self.var_ble_rssi_warn.set("")
        self._show_version_panel()
        self.lbl_log_hint.config(text=f"半成品测试 → {port}")
        self._log(f"打开串口 {port} …", "info")

        threading.Thread(target=self._run_serial, args=(port,), daemon=True).start()

    def _format_hex(self, text: str) -> str:
        """智能格式化：前缀文本保持原样，hex 数据部分每2位加空格"""
        s = text.strip()
        if ":" in s:
            prefix, hex_part = s.split(":", 1)
            hex_part = hex_part.replace(" ", "")
            formatted = " ".join(hex_part[i:i+2] for i in range(0, len(hex_part), 2))
            return f"{prefix}: {formatted}"
        else:
            clean = s.replace(" ", "")
            return " ".join(clean[i:i+2] for i in range(0, len(clean), 2))

    def _on_hex_entry_focus_out(self, var, event=None):
        """输入框失焦时自动格式化：去除空格→仅保留 hex 字符→大写→每2位加空格"""
        raw = var.get()
        # 去除空格
        stripped = raw.replace(" ", "")
        # 仅保留 0-9 a-f A-F，转大写
        hex_only = "".join(c.upper() for c in stripped if c.isalnum() and c.upper() in "0123456789ABCDEF")
        # 每2位加空格
        formatted = " ".join(hex_only[i:i+2] for i in range(0, len(hex_only), 2))
        if formatted != raw:
            var.set(formatted)

    @staticmethod
    def _hex_to_int8(hex_str: str) -> int:
        """hex 字符串 → int8（有符号 -128~127）"""
        val = int(hex_str, 16)
        return val - 256 if val > 127 else val

    def _handle_version_line(self, text: str):
        """检测 S R rf : 23 + 数分微站号 版本行，提取倒数3/4/5/6字节
        返回 (ver_name, ver_str) 或 None"""
        if not self._is_version_line(text):
            return None

        # 提取 hex 部分
        if ":" in text:
            _, hp = text.split(":", 1)
        else:
            hp = text
        hb = hp.strip().split()
        if len(hb) < 6:
            return None

        # 倒数第6 = index，倒数第5/4/3 = version
        try:
            idx = int(hb[-6], 16)
        except ValueError:
            return None
        ver_bytes = hb[-6:-2]  # 4 字节（含索引）
        ver_str = " ".join(ver_bytes)

        idx_map = {0: "主机版本00", 1: "从机版本01", 2: "从机版本02"}
        name = idx_map.get(idx)
        if name is None:
            return None
        return name, ver_str

    def _build_line_segments(self, text: str):
        """将串口行分解为 (text, tag) 片段，同时返回 (segments, rssi_int8, ble_rssi_int8)"""
        segments = []
        rssi_int8 = None
        ble_rssi_int8 = None

        if ":" in text:
            prefix, hex_part = text.split(":", 1)
            segments.append((prefix + ": ", "plain"))
        else:
            hex_part = text

        hex_bytes = hex_part.strip().split()
        if not hex_bytes:
            segments.append((hex_part, "plain"))
            return segments, rssi_int8, ble_rssi_int8

        # 获取匹配关键词及对应 tag（去空格大写）
        kw_map = [
            (self.var_ble_tag.get().replace(" ", "").upper(),    "match_ble"),
            (self.var_station_id.get().replace(" ", "").upper(), "match_station"),
        ]

        hex_clean = "".join(hex_bytes).upper()
        byte_tags = {}   # byte_index → tag

        for kw, kw_tag in kw_map:
            if not kw:
                continue
            pos = 0
            kw_len = len(kw)
            while True:
                idx = hex_clean.find(kw, pos)
                if idx == -1:
                    break
                byte_start = idx // 2
                byte_end = (idx + kw_len - 1) // 2 + 1
                for bi in range(byte_start, byte_end):
                    if bi < len(hex_bytes):
                        byte_tags[bi] = kw_tag
                pos = idx + 1

        # 蓝牙标签匹配后第7个字节标红
        ble_kw = self.var_ble_tag.get().replace(" ", "").upper()
        if ble_kw:
            pos = 0
            ble_kw_len = len(ble_kw)
            while True:
                idx = hex_clean.find(ble_kw, pos)
                if idx == -1:
                    break
                match_byte_end = (idx + ble_kw_len - 1) // 2 + 1
                target = match_byte_end + 5  # 第6个字节（0-index）
                if target < len(hex_bytes):
                    byte_tags[target] = "ble_7th"
                pos = idx + 1

        # 最后一个字节
        last_idx = len(hex_bytes) - 1 if hex_bytes else None
        if last_idx is not None and last_idx not in byte_tags:
            byte_tags[last_idx] = "last"

        # 提取信号强度 int8 值
        # 基站信号强度 = 最后字节 → int8
        if hex_bytes:
            try:
                rssi_int8 = self._hex_to_int8(hex_bytes[-1])
            except (ValueError, IndexError):
                pass
        # 蓝牙标签信号强度 = ble_7th 标记字节 → int8
        for bi, t in byte_tags.items():
            if t == "ble_7th":
                try:
                    ble_rssi_int8 = self._hex_to_int8(hex_bytes[bi])
                except (ValueError, IndexError):
                    pass
                break

        for i, byte_val in enumerate(hex_bytes):
            tag = byte_tags.get(i, "plain")
            sep = " " if i < len(hex_bytes) - 1 else ""
            segments.append((byte_val + sep, tag))

        return segments, rssi_int8, ble_rssi_int8

    def _log_segments(self, segments):
        """写入带多标签的日志行"""
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.config(state="normal")
        self.log_box.insert("end", f"[{ts}] ", "ts")
        for text, tag in segments:
            self.log_box.insert("end", text, tag)
        self.log_box.insert("end", "\n", "plain")
        self.log_box.see("end")
        self.log_box.config(state="disabled")

    def _is_version_line(self, text: str):
        """版本行判定：原始字符串直接匹配 S R rf : 23 + 数分微站号"""
        station = self.var_station_id.get().strip().upper()
        if not station:
            return False
        return text.upper().lstrip().startswith(f"S R RF : 23 {station}")

    def _should_display(self, text: str):
        """过滤：先看是不是版本行，不是再看双匹配（蓝牙标签+数分微站号同时出现）"""
        # 版本行：原始字符串直接匹配
        if self._is_version_line(text):
            return True
        # 非版本行 → 必须同时包含蓝牙测试标签和数分微站号
        stations = self.var_station_id.get().replace(" ", "").upper()
        bles = self.var_ble_tag.get().replace(" ", "").upper()
        if not bles or not stations:
            return False
        line_clean = text.replace(" ", "").upper()
        return bles in line_clean and stations in line_clean

    def _is_signal_line(self, text: str):
        """信号行判定：非版本行 + 双匹配，用于信号强度累加"""
        # 版本行不算信号
        if self._is_version_line(text):
            return False
        stations = self.var_station_id.get().replace(" ", "").upper()
        bles = self.var_ble_tag.get().replace(" ", "").upper()
        if not bles or not stations:
            return False
        line_clean = text.replace(" ", "").upper()
        return bles in line_clean and stations in line_clean

    def _run_serial(self, port):
        try:
            ser = serial.Serial(port, baudrate=115200, timeout=1)
            self._log(f"串口 {port} 已打开（115200bps）", "ok")

            while self.running:
                try:
                    raw = ser.readline()
                    if raw:
                        # ASCII 解码不会破坏 hex 文本数据
                        text = raw.decode("ascii", errors="replace").rstrip()
                        # 按 _should_display 过滤
                        if self._should_display(text):
                            self.log_queue.put(("line", text))
                except serial.SerialException as e:
                    self.log_queue.put(("err", f"串口错误：{e}"))
                    break
                except Exception as e:
                    self.log_queue.put(("err", f"读取异常：{e}"))
                    break

            ser.close()
            self._log(f"串口 {port} 已关闭", "info")
        except serial.SerialException as e:
            self.log_queue.put(("err", f"无法打开串口 {port}：{e}"))
        except Exception as e:
            self.log_queue.put(("err", f"串口启动失败：{e}"))
        finally:
            self.log_queue.put(("finished", None))

    def _launch(self, idx):
        prog = PROGRAMS[idx]
        exe = prog["exe"]
        if not os.path.exists(exe):
            messagebox.showerror("文件不存在", f"找不到文件：\n{exe}")
            return

        self.running = True
        self.lbl_log_hint.config(text=f"正在运行：{prog['label']}")
        self._log(f"启动 {prog['label']}：{exe}", "info")
        threading.Thread(target=self._run_process, args=(exe,), daemon=True).start()

    def _start(self):
        if self.running:
            return
        prog = PROGRAMS[self.var_prog.get()]
        exe = prog["exe"]
        if not os.path.exists(exe):
            messagebox.showerror("文件不存在", f"找不到文件：\n{exe}")
            return

        self.running = True
        self.btn_start._set_state("disabled")
        self.btn_stop._set_state("normal")
        self.lbl_status.config(text="● 运行中", fg="#E67E22")
        self.lbl_log_hint.config(text=f"正在运行：{prog['label']}")

        self._log(f"启动：{exe}", "info")

        t = threading.Thread(target=self._run_process, args=(exe,), daemon=True)
        t.start()

    def _stop(self):
        """停止测试：终止进程/串口，清理 UI 状态"""
        if self.running:
            self.running = False
            if self.proc and self.proc.poll() is None:
                self.proc.terminate()
            self._log("已发送停止信号", "warn")
        # 取消成品测试超时定时器
        tid = getattr(self, "_full_test_timeout_id", None)
        if tid:
            self.after_cancel(tid)
            self._full_test_timeout_id = None
        # 无论 running 是否已为 False，都执行清理（处理自动完成/超时的残留状态）
        self._stop_progress()
        self._hide_version_panel()
        self._unlock_config()
        # 重置信号显示
        self.var_rssi_disp.set("— dBm"); self.var_ble_rssi_disp.set("— dBm")
        self._rssi_count = 0; self._ble_rssi_count = 0
        self.var_rssi_count.set(""); self.var_ble_rssi_count.set("")
        self.var_rssi_warn.set(""); self.var_ble_rssi_warn.set("")
        self._rssi_values = []; self._ble_rssi_values = []
        self._test_mode = ""
        # 清除重试写入状态
        self._pending_write_data = None
        self.btn_retry.pack_forget()

    def _clear_log(self):
        self.log_box.config(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.config(state="disabled")

    # ─────────────────────────────────────────
    # 进度条
    # ─────────────────────────────────────────
    SEMI_DURATION = 240  # 半成品测试固定 4 分钟

    def _start_progress(self, duration=None):
        """启动进度条（duration=None 为成品测试按采集包数推进，有值则按时间倒计时）"""
        if duration is not None:
            # 半成品测试：时间倒计时
            self._progress_duration = duration
            self.progress_bar["value"] = 0
            self.var_progress_pct.set("0%")
            m, s = divmod(duration, 60)
            self.lbl_progress_time.config(text=f"{m}:{s:02d}")
        else:
            # 成品测试：按采集包数推进
            self._progress_duration = None
            self.progress_bar["value"] = 0
            self.var_progress_pct.set("0%")
            self.lbl_progress_time.config(text="—")
        self._progress_start = datetime.now()
        self._tick_progress()

    def _tick_progress(self):
        if not self.running:
            return
        duration = getattr(self, "_progress_duration", None)
        if duration is not None:
            # 半成品测试：时间驱动
            elapsed = (datetime.now() - self._progress_start).total_seconds()
            pct = min(elapsed / duration * 100, 100)
            self.progress_bar["value"] = pct
            self.var_progress_pct.set(f"{int(pct)}%")
            remain = max(duration - elapsed, 0)
            m, s = divmod(int(remain), 60)
            self.lbl_progress_time.config(text=f"{m}:{s:02d}")
            if pct < 100:
                self._progress_id = self.after(500, self._tick_progress)
            else:
                if not self._versions_checked:
                    self._versions_checked = True
                    self.var_ver_status.set("4分钟未采集到")
                    self.lbl_ver_status.config(fg=TEXT_ERR)
                    self._log("4分钟超时，版本未采集完成", "err")
                    self.running = False
                    self._check_rssi_thresholds()
        else:
            # 成品测试：采集包数驱动
            try:
                total = int(self.var_ble_packet_count.get())
            except (ValueError, TypeError):
                total = 10
            pct = min(self._ble_rssi_count / total * 100, 100)
            self.progress_bar["value"] = pct
            self.var_progress_pct.set(f"{int(pct)}%")
            self.lbl_progress_time.config(text=f"{self._ble_rssi_count}/{total}")
            self._progress_id = self.after(500, self._tick_progress)

    def _check_rssi_thresholds(self):
        """计算信号中位数并与阈值比较"""
        # 基站信号
        thresh = int(self.var_rssi_thresh.get())
        if self._rssi_values:
            self._rssi_values.sort()
            n = len(self._rssi_values)
            median = self._rssi_values[n // 2] if n % 2 == 1 else \
                     (self._rssi_values[n // 2 - 1] + self._rssi_values[n // 2]) // 2
            if median < thresh:
                self.var_rssi_warn.set(f"基站信号弱（中位数 {median:+d} < 阈值 {thresh:+d}）")
                self._log(f"基站信号弱：中位数 {median:+d} dBm < 阈值 {thresh:+d} dBm", "warn")
        # 蓝牙标签信号
        tag_thresh = int(self.var_tag_rssi_thresh.get())
        if self._ble_rssi_values:
            self._ble_rssi_values.sort()
            n = len(self._ble_rssi_values)
            median = self._ble_rssi_values[n // 2] if n % 2 == 1 else \
                     (self._ble_rssi_values[n // 2 - 1] + self._ble_rssi_values[n // 2]) // 2
            if median < tag_thresh:
                self.var_ble_rssi_warn.set(f"与蓝牙标签信号弱（中位数 {median:+d} < 阈值 {tag_thresh:+d}）")
                self._log(f"与蓝牙标签信号弱：中位数 {median:+d} dBm < 阈值 {tag_thresh:+d} dBm", "warn")

    # ─────────────────────────────────────────
    # Excel 写入
    # ─────────────────────────────────────────
    def _write_to_excel(self, mobile_bt="", timeout_remarks=None):
        """将测试结果写入 Excel 的第一个工作表
        timeout_remarks: 超时模式下只写编号和备注，跳过信号强度列"""
        filepath = getattr(self, "_excel_fullpath", "")
        if not filepath or not os.path.isfile(filepath):
            self._log("未找到 Excel 文件，跳过写入", "warn")
            return
        if not HAS_OPENPYXL:
            self._log("openpyxl 未安装，跳过写入", "warn")
            return

        try:
            wb = load_workbook(filepath)
        except Exception:
            wb = Workbook()

        sheet = wb.active  # 第一个工作表

        # 写入表头（如果第 1 行为空）
        headers = [
            "编号", "与基站信号强度", "与蓝牙标签信号强度",
            "蓝牙信号强度", "备注",
        ]
        existing = sheet.cell(row=1, column=1).value
        if not existing:
            for ci, h in enumerate(headers, 1):
                sheet.cell(row=1, column=ci, value=h)

        # ── 编号：取数分微站号（去空格） ──
        sid = self.var_station_id.get().replace(" ", "")

        # 辅助：判断一行是否完全为空（5 列全 None）
        def _row_empty(r):
            for c in range(1, 6):
                if sheet.cell(row=r, column=c).value is not None:
                    return False
            return True

        # 辅助：判断一行是否有任何数据
        def _row_has_data(r):
            for c in range(1, 6):
                if sheet.cell(row=r, column=c).value is not None:
                    return True
            return False

        # ── 第一遍：收集要删除的行号 + 记录覆盖位置 ──
        to_delete = set()
        overwrite_row = None
        row_idx = 2
        while True:
            # 先判断当前行是否完全为空
            if _row_empty(row_idx):
                # 再确认下方是否还有数据，如果有说明这是夹在中间的空行
                has_below = False
                check_r = row_idx + 1
                while True:
                    if check_r > sheet.max_row:
                        break
                    if _row_has_data(check_r):
                        has_below = True
                        break
                    check_r += 1
                if has_below:
                    # 夹在中间的空行，标记删除
                    to_delete.add(row_idx)
                    row_idx += 1
                    continue
                else:
                    # 已经是末尾连续空行，退出扫描
                    break

            # 有数据的行：检查是否重复编号
            val_a = sheet.cell(row=row_idx, column=1).value
            if val_a is not None and str(val_a).replace(" ", "") == sid:
                # 匹配到重复编号，记住位置，不删除（稍后覆盖写入）
                overwrite_row = row_idx
            row_idx += 1

        # ── 第二遍：从后往前删除空行，避免索引变化 ──
        for r in sorted(to_delete, reverse=True):
            sheet.delete_rows(r)

        # 如果匹配行在被删除的空行之后，需要修正行号
        if overwrite_row is not None:
            deleted_before = sum(1 for d in to_delete if d < overwrite_row)
            overwrite_row -= deleted_before

        # ── 第三遍：确定写入行 ──
        if overwrite_row is not None:
            # 覆盖写入已有行
            row_idx = overwrite_row
            is_overwrite = True
            # 先清空该行旧数据
            for ci in range(1, 6):
                sheet.cell(row=row_idx, column=ci, value=None)
        else:
            # 找到末尾空行追加
            row_idx = 2
            while not _row_empty(row_idx):
                row_idx += 1
            is_overwrite = False

        row_data = [sid if sid else row_idx, "", "", "", ""]
        sheet.cell(row=row_idx, column=1, value=row_data[0])

        if timeout_remarks is not None:
            # 超时模式：只写编号 + 备注，跳过信号强度
            row_data[4] = timeout_remarks
            sheet.cell(row=row_idx, column=5, value=timeout_remarks)
        else:
            # ── 与基站信号强度（均值，取整）──
            if self._rssi_values:
                avg_rssi = int(sum(self._rssi_values) / len(self._rssi_values))
                row_data[1] = str(avg_rssi)
                sheet.cell(row=row_idx, column=2, value=avg_rssi)

            # ── 与蓝牙标签信号强度（均值，取整）──
            if self._ble_rssi_values:
                avg_ble = int(sum(self._ble_rssi_values) / len(self._ble_rssi_values))
                row_data[2] = str(avg_ble)
                sheet.cell(row=row_idx, column=3, value=avg_ble)

            # ── 蓝牙信号强度（手动输入，成品测试才有）──
            if mobile_bt:
                row_data[3] = str(mobile_bt)
                try:
                    sheet.cell(row=row_idx, column=4, value=int(mobile_bt))
                except ValueError:
                    sheet.cell(row=row_idx, column=4, value=mobile_bt)

            # ── 备注 ──
            remarks = []
            if self.var_rssi_warn.get():
                remarks.append("基站信号弱")
            if self.var_ble_rssi_warn.get():
                remarks.append("与蓝牙标签信号弱")
            if self._test_mode == "semi" and self._versions_checked:
                if self.var_ver_status.get() == "4分钟未采集到":
                    remarks.append("4分钟未采集到版本")
                elif self.var_ver_status.get() == "版本异常":
                    remarks.append("版本异常")
            row_data[4] = "; ".join(remarks)
            sheet.cell(row=row_idx, column=5, value=row_data[4])

        # ── 自动列宽（最低 18）──
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(headers) + 1):
            max_len = 0
            for r in range(1, row_idx + 1):
                val = sheet.cell(row=r, column=col_idx).value
                if val is not None:
                    w = sum(2 if ord(c) > 127 else 1 for c in str(val))
                    max_len = max(max_len, w)
            width = max(max_len + 2, 18)
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        # ── 全部靠右对齐 ──
        from openpyxl.styles import Alignment
        for r in range(1, row_idx + 1):
            for c in range(1, len(headers) + 1):
                sheet.cell(row=r, column=c).alignment = Alignment(horizontal="right")

        action_text = "覆盖写入" if is_overwrite else "写入"
        # ── 构建日志展示内容 ──
        log_lines = [
            "────────────── 写入内容 ──────────────",
            f"行号：{row_idx}" + ("（覆盖已有数据）" if is_overwrite else ""),
            f"编号：{row_data[0]}",
            f"与基站信号强度：{row_data[1] or '—'}",
            f"与蓝牙标签信号强度：{row_data[2] or '—'}",
            f"蓝牙信号强度：{row_data[3] or '—'}",
            f"备注：{row_data[4] or '—'}",
        ]

        try:
            wb.save(filepath)
            self._log(f"数据已{action_text} {os.path.basename(filepath)} 第 {row_idx} 行", "ok")
            self.btn_retry.pack_forget()
        except PermissionError:
            # 文件被占用（Excel 打开中）
            content = "\n".join(log_lines)
            self._log(content, "warn")
            self._log("请关闭工作表后点击下方 [重新写入] 按钮", "err")
            self._pending_write_data = (filepath, row_data, row_idx, headers, wb)
            self.btn_retry.pack(before=self.progress_frame, fill="x", padx=4, pady=(4, 0))
        except Exception as e:
            content = "\n".join(log_lines)
            self._log(content, "warn")
            self._log(f"写入失败：{e}，请关闭工作表后点击下方 [重新写入] 按钮", "err")
            self._pending_write_data = (filepath, row_data, row_idx, headers, wb)
            self.btn_retry.pack(before=self.progress_frame, fill="x", padx=4, pady=(4, 0))

    def _retry_write(self):
        """重新尝试写入 Excel"""
        if not self._pending_write_data:
            self._log("无待写入数据", "warn")
            self.btn_retry.pack_forget()
            return
        filepath, row_data, row_idx, headers, wb = self._pending_write_data
        self._log("正在重新写入…", "info")
        try:
            wb.save(filepath)
            self._log(f"数据已写入 {os.path.basename(filepath)} 第 {row_idx} 行", "ok")
            self._pending_write_data = None
            self.btn_retry.pack_forget()
        except PermissionError:
            self._log("文件仍被占用，请先关闭工作表", "err")
        except Exception as e:
            self._log(f"写入失败：{e}", "err")

    # ─────────────────────────────────────────
    # 成品测试完成
    # ─────────────────────────────────────────
    def _on_full_test_done(self):
        """成品测试完成：采集到足够包数 → 信号评估 → 弹出手机蓝牙信号输入"""
        if self._test_mode != "full":
            return  # 防止重复调用
        # 取消进度条定时器，保留 100% 显示
        pid = getattr(self, "_progress_id", None)
        if pid is not None:
            self.after_cancel(pid)
            self._progress_id = None
        self.progress_bar["value"] = 100
        self.var_progress_pct.set("100%")
        self._check_rssi_thresholds()
        self._log("成品测试采集完成", "ok")
        # 关闭串口（串口线程通过 running=False 退出）
        self.running = False
        # 弹出手机蓝牙信号输入
        self.after(300, self._ask_mobile_bt)

    def _ask_mobile_bt(self):
        """弹出输入框：手机蓝牙信号 → 写入 Excel → 解锁配置"""
        from tkinter import simpledialog
        val = simpledialog.askstring(
            "手机蓝牙信号",
            "请输入手机蓝牙信号强度（dBm）：",
            parent=self,
        )
        mobile_bt = ""
        if val is not None:
            self._log(f"手机蓝牙信号：{val} dBm", "ok")
            mobile_bt = val
        else:
            self._log("手机蓝牙信号：未输入", "warn")
        # 写入 Excel
        self._write_to_excel(mobile_bt)
        # ── 解锁配置 ──
        self._unlock_config()

    def _on_full_test_timeout(self):
        """成品测试超时：N秒内未收到信号数据 → 记录超时 → 停止 → 解锁配置"""
        self._full_test_timeout_id = None
        # 关闭串口
        self.running = False
        # 取消进度条
        pid = getattr(self, "_progress_id", None)
        if pid is not None:
            self.after_cancel(pid)
            self._progress_id = None
        self.progress_bar["value"] = 0
        self.var_progress_pct.set("超时")
        self.lbl_progress_time.config(text="超时")
        try:
            timeout_sec = int(self.var_ble_timeout.get())
        except (ValueError, TypeError):
            timeout_sec = 20
        self._log(f"{timeout_sec}秒内未收到蓝牙标签或微站数据 — 记录超时", "err")
        # 写入 Excel：仅编号 + 备注
        self._write_to_excel(timeout_remarks="蓝牙标签或微站没有数据上报")
        # ── 解锁配置 ──
        self._unlock_config()

    def _stop_progress(self):
        """取消进度条定时器并复位"""
        pid = getattr(self, "_progress_id", None)
        if pid is not None:
            self.after_cancel(pid)
            self._progress_id = None
        self.progress_bar["value"] = 0
        self.var_progress_pct.set("0%")
        self.lbl_progress_time.config(text="")

    # ─────────────────────────────────────────
    # 进程管理
    # ─────────────────────────────────────────
    def _run_process(self, exe):
        try:
            self.proc = subprocess.Popen(
                [exe],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                cwd=os.path.dirname(exe),
                encoding="gbk",
                errors="replace",
                creationflags=subprocess.CREATE_NO_WINDOW,
            )
            for line in iter(self.proc.stdout.readline, ""):
                self.log_queue.put(("line", line.rstrip()))
            self.proc.stdout.close()
            self.proc.wait()
            code = self.proc.returncode
            if code == 0:
                self.log_queue.put(("done_ok", f"进程正常退出（返回码 {code}）"))
            else:
                self.log_queue.put(("done_err", f"进程退出（返回码 {code}）"))
        except Exception as e:
            self.log_queue.put(("err", f"启动失败：{e}"))
        finally:
            self.log_queue.put(("finished", None))

    # ─────────────────────────────────────────
    # 日志队列轮询（主线程）
    # ─────────────────────────────────────────
    def _poll_log(self):
        try:
            while True:
                tag, msg = self.log_queue.get_nowait()
                if tag == "line":
                    # ── 半成品测试：检测版本行 ──
                    if self._test_mode != "full":
                        vinfo = self._handle_version_line(msg)
                        if vinfo:
                            name, ver = vinfo
                            self._ver_vars[name].set(ver)
                            self._collected_versions[name] = ver
                            self._check_versions_complete()
                    segments, rssi, ble_rssi = self._build_line_segments(msg)
                    self._log_segments(segments)
                    # 仅信号行（蓝牙标签+数分微站号同时匹配）参与信号计算，版本行不参与
                    if self._is_signal_line(msg):
                        # 成品测试：首次收到信号数据 → 取消超时定时器
                        if self._test_mode == "full":
                            tid = getattr(self, "_full_test_timeout_id", None)
                            if tid:
                                self.after_cancel(tid)
                                self._full_test_timeout_id = None
                        if rssi is not None:
                            self._rssi_count += 1
                            self._rssi_values.append(rssi)
                            self.var_rssi_disp.set(f"{rssi:+d} dBm")
                            self.var_rssi_count.set(f"×{self._rssi_count}")
                        if ble_rssi is not None:
                            self._ble_rssi_count += 1
                            self._ble_rssi_values.append(ble_rssi)
                            self.var_ble_rssi_disp.set(f"{ble_rssi:+d} dBm")
                            self.var_ble_rssi_count.set(f"×{self._ble_rssi_count}")
                            # 成品测试：采集到信号采集包数 → 结束（然后回到监听）
                            if self._test_mode == "full":
                                try:
                                    target = int(self.var_ble_packet_count.get())
                                except (ValueError, TypeError):
                                    target = 10
                                if self._ble_rssi_count >= target:
                                    self._on_full_test_done()
                elif tag == "done_ok":
                    self._log(msg, "ok")
                    self._on_process_done()
                elif tag == "done_err":
                    self._log(msg, "err")
                    self._on_process_done()
                elif tag == "err":
                    self._log(msg, "err")
                    self._on_process_done()
                elif tag == "finished":
                    self._on_process_done()
        except queue.Empty:
            pass
        self.after(100, self._poll_log)

    def _on_process_done(self):
        if not self.running:
            return
        self.running = False
        self.proc = None
        self._unlock_config()
        # 自动完成时仅更新提示，保留所有测试结果显示
        self.lbl_log_hint.config(text="测试结束")

    # ─────────────────────────────────────────
    # 写日志
    # ─────────────────────────────────────────
    def _log(self, msg, tag="plain"):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.config(state="normal")
        self.log_box.insert("end", f"[{ts}] ", "ts")
        self.log_box.insert("end", f"{msg}\n", tag)
        self.log_box.see("end")
        self.log_box.config(state="disabled")


if __name__ == "__main__":
    app = LauncherApp()
    app.mainloop()
